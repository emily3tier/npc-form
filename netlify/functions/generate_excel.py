import json
import base64
import io
import urllib.request
import urllib.parse
import os
import datetime
from openpyxl import load_workbook

CLIENT_ID   = 'd6dd88e1-a49e-4350-8339-f0f42c4b3b2e'
TENANT_ID   = '667afa82-1126-4a78-8f76-0918c7f2a845'
BASE_FOLDER = 'UPC Submissions Automated'
TEMPLATE_NAME = 'NPC_Form_2026_1.xlsx'

def get_access_token(refresh_token):
    data = urllib.parse.urlencode({'client_id': CLIENT_ID,'grant_type': 'refresh_token','refresh_token': refresh_token,'scope': 'Files.ReadWrite offline_access User.Read',}).encode()
    req = urllib.request.Request(f'https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token',data=data, headers={'Content-Type': 'application/x-www-form-urlencoded'})
    resp = urllib.request.urlopen(req)
    return json.loads(resp.read())['access_token']

def graph(token, method, path, body=None, content_type='application/json'):
    url = 'https://graph.microsoft.com/v1.0' + path
    data = None
    if body is not None:
        data = body if isinstance(body, (bytes, bytearray)) else json.dumps(body).encode()
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header('Authorization', 'Bearer ' + token)
    if data:
        req.add_header('Content-Type', content_type)
    try:
        resp = urllib.request.urlopen(req)
        return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        if e.code == 409:
            return {}
        raise Exception(f'Graph {e.code}: {e.read().decode()}')

def graph_download(token, path):
    url = 'https://graph.microsoft.com/v1.0' + path
    req = urllib.request.Request(url, method='GET')
    req.add_header('Authorization', 'Bearer ' + token)
    return urllib.request.urlopen(req).read()

def ensure_folder(token, parent_path, folder_name):
    if parent_path:
        encoded = '/'.join(urllib.parse.quote(s, safe='') for s in parent_path.split('/'))
        url = f'/me/drive/root:/{encoded}:/children'
    else:
        url = '/me/drive/root/children'
    graph(token, 'POST', url, {'name': folder_name, 'folder': {}})

def upload_file(token, folder_path, file_name, data):
    encoded = '/'.join(urllib.parse.quote(s, safe='') for s in (folder_path + '/' + file_name).split('/'))
    graph(token, 'PUT', f'/me/drive/root:/{encoded}:/content', data, 'application/octet-stream')

def get_template(token):
    return graph_download(token, f'/me/drive/root:/{urllib.parse.quote(TEMPLATE_NAME, safe="")}:/content')

def build_excel(template_bytes, form_data, products):
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb['New Product Coding Form']
    ws['B4'] = form_data.get('fromName', '')
    ws['D4'] = form_data.get('clientName', '')
    ws['B5'] = form_data.get('email', '')
    ws['D5'] = form_data.get('ownProduct', 'Yes')
    coding = form_data.get('codingOption', 'Code Immediately')
    ws['A7'] = 'X' if coding == 'Code Immediately' else ''
    ws['A8'] = 'X' if coding == 'Delay until Sales' else ''
    ws['A9'] = 'X' if coding == 'Code by Saturday Date' else ''
    if coding == 'Code by Saturday Date' and form_data.get('saturdayDate'):
        ws['C9'] = form_data['saturdayDate']
    ws['E10'] = form_data.get('containerType', '')
    ws['F10'] = form_data.get('containerMaterial', '')
    addl = form_data.get('additionalInfo', '')
    for i, p in enumerate(products):
        row = 12 + i
        ws[f'A{row}'] = p.get('upc', '')
        ws[f'B{row}'] = p.get('asin', '')
        ws[f'C{row}'] = p.get('costco', '')
        ws[f'D{row}'] = p.get('name', '') + (' | ' + addl if addl else '')
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def handler(event, context):
    cors = {'Access-Control-Allow-Origin': '*','Access-Control-Allow-Headers': 'Content-Type','Access-Control-Allow-Methods': 'POST, OPTIONS'}
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': cors, 'body': ''}
    try:
        refresh_token = os.environ.get('MS_REFRESH_TOKEN', '')
        if not refresh_token:
            raise Exception('MS_REFRESH_TOKEN not set')
        body = json.loads(event['body'])
        form_data = body.get('formData', {})
        products = body.get('products', [])
        files = body.get('files', [])
        token = get_access_token(refresh_token)
        ensure_folder(token, '', BASE_FOLDER)
        for p in products:
            ensure_folder(token, BASE_FOLDER, p.get('folderName', ''))
        for f in files:
            upload_file(token, BASE_FOLDER + '/' + f.get('folderName', ''), f.get('fileName', ''), base64.b64decode(f.get('data', '')))
        template_bytes = get_template(token)
        excel_bytes = build_excel(template_bytes, form_data, products)
        client_safe = ''.join(c if c.isalnum() else '_' for c in form_data.get('clientName', 'client'))
        excel_name = f'NPC_Form_{client_safe}_{datetime.date.today().isoformat()}.xlsx'
        upload_file(token, BASE_FOLDER, excel_name, excel_bytes)
        return {'statusCode': 200, 'headers': cors, 'body': json.dumps({'excel': base64.b64encode(excel_bytes).decode(), 'excelName': excel_name})}
    except Exception as e:
        return {'statusCode': 500, 'headers': cors, 'body': json.dumps({'error': str(e)})}
